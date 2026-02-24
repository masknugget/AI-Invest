"""
Excel 日志导出器
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional
import logging

from app.services.logging.models import ExportResult

logger = logging.getLogger("webapi")


class ExcelExporter:
    """Excel 日志导出器"""
    
    def __init__(self, export_dir: str = "exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
    
    async def export(
        self,
        logs: List[Dict[str, Any]],
        filename: Optional[str] = None,
        sheet_name: str = "Logs",
        columns: Optional[List[str]] = None,
        auto_width: bool = True
    ) -> ExportResult:
        """
        导出日志到 Excel
        
        Args:
            logs: 日志列表
            filename: 文件名（不含扩展名）
            sheet_name: 工作表名称
            columns: 指定导出的列
            auto_width: 自动调整列宽
        """
        if not logs:
            return ExportResult(
                success=True,
                record_count=0,
                message="No logs to export"
            )
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return ExportResult(
                success=False,
                record_count=0,
                message="openpyxl is required for Excel export. Install with: pip install openpyxl"
            )
        
        # 生成文件名
        if not filename:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"logs_export_{timestamp}"
        
        file_path = self.export_dir / f"{filename}.xlsx"
        
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # 确定列
            if not columns:
                all_keys = set()
                for log in logs:
                    all_keys.update(log.keys())
                # 优先显示常用字段
                priority_fields = ["timestamp", "level", "log_type", "user_id", "action", "message"]
                columns = priority_fields + sorted(all_keys - set(priority_fields))
            
            # 写入表头
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入数据
            for row_idx, log in enumerate(logs, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    value = log.get(col_name, "")
                    
                    # 处理复杂类型
                    if isinstance(value, dict):
                        value = str(value)
                    elif isinstance(value, list):
                        value = ", ".join(str(v) for v in value)
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # 交替行颜色
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # 自动调整列宽
            if auto_width:
                for col_idx, col_name in enumerate(columns, 1):
                    max_length = len(col_name)
                    for log in logs:
                        value = str(log.get(col_name, ""))
                        max_length = max(max_length, min(len(value), 100))  # 限制最大宽度
                    
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
            
            # 冻结首行
            ws.freeze_panes = "A2"
            
            # 保存
            wb.save(file_path)
            
            file_size = file_path.stat().st_size
            
            return ExportResult(
                success=True,
                file_path=str(file_path),
                file_name=file_path.name,
                file_size_bytes=file_size,
                record_count=len(logs),
                format="xlsx",
                created_at=datetime.utcnow()
            )
            
        except Exception as e:
            logger.error(f"Excel导出失败: {e}")
            return ExportResult(
                success=False,
                record_count=0,
                message=f"Export failed: {str(e)}"
            )
    
    async def export_multiple_sheets(
        self,
        data: Dict[str, List[Dict[str, Any]]],
        filename: Optional[str] = None
    ) -> ExportResult:
        """
        导出多个工作表
        
        Args:
            data: {sheet_name: logs}
            filename: 文件名
        """
        if not data:
            return ExportResult(
                success=True,
                record_count=0,
                message="No data to export"
            )
        
        try:
            import openpyxl
        except ImportError:
            return ExportResult(
                success=False,
                record_count=0,
                message="openpyxl is required for Excel export"
            )
        
        if not filename:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"logs_export_{timestamp}"
        
        file_path = self.export_dir / f"{filename}.xlsx"
        
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # 删除默认工作表
            
            total_records = 0
            
            for sheet_name, logs in data.items():
                ws = wb.create_sheet(title=sheet_name[:31])  # Excel限制31字符
                
                if not logs:
                    continue
                
                # 确定列
                all_keys = set()
                for log in logs:
                    all_keys.update(log.keys())
                columns = sorted(all_keys)
                
                # 写入表头
                for col_idx, col_name in enumerate(columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                
                # 写入数据
                for row_idx, log in enumerate(logs, 2):
                    for col_idx, col_name in enumerate(columns, 1):
                        value = log.get(col_name, "")
                        if isinstance(value, (dict, list)):
                            value = str(value)
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                total_records += len(logs)
            
            wb.save(file_path)
            
            file_size = file_path.stat().st_size
            
            return ExportResult(
                success=True,
                file_path=str(file_path),
                file_name=file_path.name,
                file_size_bytes=file_size,
                record_count=total_records,
                format="xlsx"
            )
            
        except Exception as e:
            logger.error(f"Excel多工作表导出失败: {e}")
            return ExportResult(
                success=False,
                record_count=0,
                message=f"Export failed: {str(e)}"
            )
    
    async def delete_export(self, filename: str) -> bool:
        """删除导出文件"""
        file_path = self.export_dir / filename
        if not file_path.suffix == '.xlsx':
            file_path = file_path.with_suffix('.xlsx')
        
        try:
            if file_path.exists():
                file_path.unlink()
                return True
        except Exception as e:
            logger.error(f"删除导出文件失败: {e}")
        return False
    
    def list_exports(self) -> List[Dict[str, Any]]:
        """列出所有导出文件"""
        exports = []
        try:
            for file_path in self.export_dir.glob("*.xlsx"):
                stat = file_path.stat()
                exports.append({
                    "filename": file_path.name,
                    "size_bytes": stat.st_size,
                    "created_at": datetime.fromtimestamp(stat.st_mtime),
                    "path": str(file_path)
                })
        except Exception as e:
            logger.error(f"列出导出文件失败: {e}")
        
        return sorted(exports, key=lambda x: x["created_at"], reverse=True)
